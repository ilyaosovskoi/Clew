"""Office task — fill an existing .xlsx template with data.

Category: office
Section: office
Difficulty: medium

Starting tree: ``template.xlsx`` with a sheet "Sheet1" that has a header
row (Name, Score) but no data rows. The agent must:
1. Open the existing template.
2. Add 3 data rows: Alice/95, Bob/87, Carol/92.
3. Save as ``filled.xlsx``.

Pass criteria: filled.xlsx exists, has the 3 rows with correct values.
"""

from __future__ import annotations

from pathlib import Path

from ..._base import (
    Difficulty,
    EvaluationReport,
    Section,
    TaskSpec,
)


def setup(workspace: str) -> None:
    root = Path(workspace)
    root.mkdir(parents=True, exist_ok=True)
    # Build a tiny template workbook. If openpyxl isn't installed
    # (e.g. minimal CI environment without the office deps), write a
    # placeholder file so the rest of the task's evaluation can still
    # run — the evaluator's openpyxl-dependent checks will simply fail
    # rather than crashing setup().
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Score"])
        wb.save(str(root / "template.xlsx"))
    except ImportError:
        (root / "template.xlsx").write_text(
            "openpyxl not installed — placeholder", encoding="utf-8"
        )


def evaluate(workspace: str, agent_output: str, tool_calls: list) -> EvaluationReport:
    root = Path(workspace)
    filled = root / "filled.xlsx"
    criteria = []

    c1 = filled.is_file()
    criteria.append({"name": "filled.xlsx exists", "passed": c1})

    # 2. Has the 3 expected rows.
    c2 = False
    if c1:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(filled), data_only=False)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            text_blob = "\n".join(
                " ".join(str(c) if c is not None else "" for c in row)
                for row in rows
            )
            c2 = (
                "Alice" in text_blob and "95" in text_blob
                and "Bob" in text_blob and "87" in text_blob
                and "Carol" in text_blob and "92" in text_blob
            )
        except Exception:
            c2 = False
    criteria.append({
        "name": "3 data rows present (Alice/95, Bob/87, Carol/92)",
        "passed": c2,
    })

    # 3. Header row preserved.
    c3 = False
    if c1:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(filled), data_only=False)
            ws = wb.active
            first_row = next(ws.iter_rows(values_only=True), None)
            if first_row:
                c3 = "Name" in [str(c) for c in first_row] and "Score" in [str(c) for c in first_row]
        except Exception:
            c3 = False
    criteria.append({"name": "header row (Name, Score) preserved", "passed": c3})

    mandatory = [c["passed"] for c in criteria[:3]]
    passed = all(mandatory)
    return EvaluationReport(
        passed=passed,
        reason="all criteria met" if passed else "criteria failed",
        checked_criteria=criteria,
    )


def build() -> TaskSpec:
    return TaskSpec(
        id="office_fill_xlsx_template",
        section=Section.OFFICE,
        difficulty=Difficulty.MEDIUM,
        description="Fill template.xlsx with 3 rows of (Name, Score) data, save as filled.xlsx.",
        prompt=(
            "Open the existing template.xlsx in the workspace root. It "
            "has a Sheet1 with a header row (Name, Score) but no data. "
            "Add these 3 data rows: Alice/95, Bob/87, Carol/92. Save the "
            "result as a NEW file called filled.xlsx (don't overwrite "
            "template.xlsx). Use the office_* tools."
        ),
        setup=setup,
        evaluate=evaluate,
        expected_duration_s=30.0,
        tags=["office", "xlsx", "template_fill"],
        min_iterations=6,
        max_time_s=240.0,
    )
