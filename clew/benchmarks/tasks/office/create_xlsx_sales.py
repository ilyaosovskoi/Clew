"""Office task — create a .xlsx with a sheet of sales data + a SUM formula.

Category: office
Section: office
Difficulty: medium

The agent must produce ``sales.xlsx`` containing:
- A sheet named "Q1"
- Headers: Month, Revenue, Cost
- 3 data rows: Jan/1000/600, Feb/1200/700, Mar/1500/800
- A formula in cell E1 (or any cell) computing SUM of Revenue.

Pass criteria: file exists, opens as valid .xlsx, sheet "Q1" exists,
the 3 data rows are present, and a formula referencing SUM exists.
"""

from __future__ import annotations

from pathlib import Path

from ..._base import (
    Difficulty,
    EvaluationReport,
    Section,
    TaskSpec,
)


def setup(workspace: str) -> None:
    root = Path(workspace)
    root.mkdir(parents=True, exist_ok=True)
    (root / "README.md").write_text(
        "# Sales workspace\n\nCreate sales.xlsx here.\n",
        encoding="utf-8",
    )


def evaluate(workspace: str, agent_output: str, tool_calls: list) -> EvaluationReport:
    root = Path(workspace)
    xlsx = root / "sales.xlsx"
    criteria = []

    c1 = xlsx.is_file()
    criteria.append({"name": "sales.xlsx exists", "passed": c1})

    c2 = False
    sheet_names: list = []
    if c1:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(xlsx), data_only=False)
            sheet_names = wb.sheetnames
            c2 = "Q1" in sheet_names
        except Exception:
            c2 = False
    criteria.append({"name": "sheet 'Q1' exists", "passed": c2})

    # 3. Headers + 3 data rows present.
    c3 = False
    if c2:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(xlsx), data_only=False)
            ws = wb["Q1"]
            rows = list(ws.iter_rows(values_only=True))
            # Look for the expected data.
            text_blob = "\n".join(
                " ".join(str(c) if c is not None else "" for c in row)
                for row in rows
            )
            c3 = (
                "Month" in text_blob and "Revenue" in text_blob and "Cost" in text_blob
                and "Jan" in text_blob and "1000" in text_blob
                and "Feb" in text_blob and "1200" in text_blob
                and "Mar" in text_blob and "1500" in text_blob
            )
        except Exception:
            c3 = False
    criteria.append({
        "name": "headers + 3 data rows (Jan/Feb/Mar) present",
        "passed": c3,
    })

    # 4. A SUM formula exists somewhere in the sheet.
    c4 = False
    if c2:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(xlsx), data_only=False)
            ws = wb["Q1"]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and "SUM" in val.upper():
                        c4 = True
                        break
                if c4:
                    break
        except Exception:
            c4 = False
    criteria.append({"name": "a SUM formula is present", "passed": c4})

    mandatory = [c["passed"] for c in criteria[:4]]
    passed = all(mandatory)
    return EvaluationReport(
        passed=passed,
        reason="all criteria met" if passed else "criteria failed",
        checked_criteria=criteria,
    )


def build() -> TaskSpec:
    return TaskSpec(
        id="office_create_xlsx_sales",
        section=Section.OFFICE,
        difficulty=Difficulty.MEDIUM,
        description="Create sales.xlsx with Q1 sheet, 3 months of data, and a SUM formula.",
        prompt=(
            "Create a new Excel file called sales.xlsx in the workspace "
            "root. It should have a sheet named 'Q1' with headers "
            "(Month, Revenue, Cost) and 3 data rows: Jan/1000/600, "
            "Feb/1200/700, Mar/1500/800. Add a SUM formula somewhere "
            "in the sheet that sums the Revenue column. Use the office_* "
            "tools (office_create, office_add_sheet, office_fill_sheet, "
            "office_set_cell)."
        ),
        setup=setup,
        evaluate=evaluate,
        expected_duration_s=40.0,
        tags=["office", "xlsx", "formulas"],
        min_iterations=8,
        max_time_s=240.0,
    )
